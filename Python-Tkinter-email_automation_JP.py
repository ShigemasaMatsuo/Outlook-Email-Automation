import os
import sys
import datetime
import openpyxl as px
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.formatting.rule import Rule
import win32com.client
import tkinter 
from tkinter import messagebox

def button_function0():
    make_dir_path = r"C:\email_automation"
    make_dir_path1 =  r"C:\email_automation\attachments"

    if os.path.isdir(make_dir_path):
        messagebox.showinfo("Message Box","[email_automation]フォルダは既に存在しています。")
        pass
    else:
        os.makedirs(make_dir_path)
        os.makedirs(make_dir_path1)
        messagebox.showinfo("Message Box","[email_automation]フォルダをCドライブ直下に作成しました!!")

def button_function1():
    check_dir_path = r"C:\email_automation"
    check_file_path = r"C:\email_automation\Email_Automation_template.xlsx"
    if os.path.isdir(check_dir_path) and os.path.isfile(check_file_path):
        messagebox.showinfo("Message Box","[Email_Automation_template.xlsx]ファイルは既に存在しています。")
        pass
    elif not os.path.isdir(check_dir_path):
        messagebox.showinfo("Message Box","[email_automation]フォルダがありません。ステップ[1]を実施してください。")
        pass
    else:
        wbx = px.Workbook()
        sheetx = wbx.active
        sheetx.title = "email contents"
        wbx.create_sheet("list")
        wbx.create_sheet("log")
        
        fill = px.styles.PatternFill(patternType='solid', fgColor='E2EFDA', bgColor='E2EFDA')
        fill1 = px.styles.PatternFill(patternType='solid', fgColor='B4C6E7', bgColor='B4C6E7')
        fill2 = px.styles.PatternFill(patternType='solid', fgColor='FFE699', bgColor='FFE699')
        fill3 = px.styles.PatternFill(patternType='solid', fgColor='FFFF00', bgColor='FFFF00')
        fill4 = px.styles.PatternFill(patternType='solid', fgColor='A6A6A6', bgColor='A6A6A6')
        lineside = Side(style="thin", color="000000")
        lineborder = Border(top=lineside, left=lineside, right=lineside, bottom=lineside)
        boldfont = Font(bold=True)
        
        wbx1 = wbx.worksheets[0]
        wbx2 = wbx.worksheets[1]
        wbx3 = wbx.worksheets[2]
        
        wbx1.cell(row=1, column=1).value = "送信元アドレス"
        wbx1.cell(row=4, column=1).value = "件名"
        wbx1.cell(row=4, column=3).value = "BCCアドレス"
        wbx1.cell(row=7, column=1).value = "本文"
        wbx1.cell(row=1, column=3).value = "CCアドレス"
        wbx1.cell(row=7, column=3).value = "署名"        
        
        wbx1.cell(row=2, column=1).fill = fill
        wbx1.cell(row=5, column=1).fill = fill
        wbx1.cell(row=8, column=1).fill = fill
        wbx1.cell(row=2, column=3).fill = fill
        wbx1.cell(row=5, column=3).fill = fill
        wbx1.cell(row=8, column=3).fill = fill
        
        wbx1.cell(column=1, row=1).border = lineborder
        wbx1.cell(column=1, row=2).border = lineborder
        wbx1.cell(column=1, row=4).border = lineborder
        wbx1.cell(column=1, row=5).border = lineborder
        wbx1.cell(column=1, row=7).border = lineborder
        wbx1.cell(column=1, row=8).border = lineborder
        wbx1.cell(column=3, row=1).border = lineborder
        wbx1.cell(column=3, row=2).border = lineborder
        wbx1.cell(column=3, row=4).border = lineborder
        wbx1.cell(column=3, row=5).border = lineborder
        wbx1.cell(column=3, row=7).border = lineborder
        wbx1.cell(column=3, row=8).border = lineborder
        
        wbx1.column_dimensions["A"].width = 55.75
        wbx1.column_dimensions["B"].width = 3.00
        wbx1.column_dimensions["C"].width = 43.58
        wbx1.row_dimensions[8].height = 300

        wbx2.cell(row=1, column=1).value = "名前"
        wbx2.cell(row=1, column=2).value = "会社名"
        wbx2.cell(row=1, column=3).value = "部署名"
        wbx2.cell(row=1, column=4).value = "Emailアドレス"
        wbx2.cell(row=1, column=5).value = "添付ファイル" 
        wbx2.cell(row=1, column=6).value = "送信対象" 
        wbx2.cell(row=1, column=8).value = "送信対象" + "\n" + "記号"
        wbx2.cell(row=2, column=8).value = "#"
        wbx2.cell(row=1, column=9).value = "添付ファイル" + "\n" + "記号" 
        wbx2.cell(row=2, column=9).value = "A"
        wbx2.cell(row=3, column=9).value = "B"
        wbx2.cell(row=4, column=9).value = "C"
        wbx2.cell(row=5, column=9).value = "D"
        wbx2.cell(row=6, column=9).value = "E"
        wbx2.cell(row=7, column=9).value = "F"
        wbx2.cell(row=1, column=10).value = "添付ファイル名" 

        wbx2.cell(row=1, column=1).fill = fill1
        wbx2.cell(row=1, column=2).fill = fill1
        wbx2.cell(row=1, column=3).fill = fill1
        wbx2.cell(row=1, column=4).fill = fill1
        wbx2.cell(row=1, column=5).fill = fill2
        wbx2.cell(row=1, column=6).fill = fill2
        wbx2.cell(row=1, column=8).fill = fill2
        wbx2.cell(row=1, column=9).fill = fill2
        wbx2.cell(row=1, column=10).fill = fill2
        wbx2.cell(row=2, column=8).fill = fill3
        wbx2.cell(row=2, column=9).fill = fill3
        wbx2.cell(row=3, column=9).fill = fill3
        wbx2.cell(row=4, column=9).fill = fill3
        wbx2.cell(row=5, column=9).fill = fill3
        wbx2.cell(row=6, column=9).fill = fill3
        wbx2.cell(row=7, column=9).fill = fill3
        wbx2.cell(row=2, column=10).fill = fill
        wbx2.cell(row=3, column=10).fill = fill
        wbx2.cell(row=4, column=10).fill = fill
        wbx2.cell(row=5, column=10).fill = fill
        wbx2.cell(row=6, column=10).fill = fill
        wbx2.cell(row=7, column=10).fill = fill

        for row in wbx2.iter_rows():
            for cell in row:
                if cell.row == 1:
                    wbx2[cell.coordinate].font = boldfont
                    cell.alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
        wbx2.cell(row=2, column=8).alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
        for i in range(2, 8):
            wbx2.cell(row=i, column=9).alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
        
        dv1 = DataValidation(
                            type="list",
                            formula1="'list'!$H$2:$H$3",
                            allow_blank=True,
                            showErrorMessage=True)
        dv1.add(f"F2:F1000")
        wbx2.add_data_validation(dv1)
        dv2 = DataValidation(
                            type="list",
                            formula1="'list'!$I$2:$I$8",
                            allow_blank=True,
                            showErrorMessage=True)
        dv2.add(f"E2:E1000")
        wbx2.add_data_validation(dv2)   
        
        for a in range(2, 1000):
            formula_rule1 = FormulaRule(formula=['$F{}=$H$2'.format(a)], fill=fill3)
            wbx2.conditional_formatting.add("F{}".format(a), formula_rule1)
            formula_rule2 = FormulaRule(formula=['$F{}=Done'.format(a)], fill=fill4)
            wbx2.conditional_formatting.add("F{}".format(a), formula_rule2)
            
        wbx2.cell(row=1, column=1).border = lineborder
        wbx2.cell(row=1, column=2).border = lineborder
        wbx2.cell(row=1, column=3).border = lineborder
        wbx2.cell(row=1, column=4).border = lineborder
        wbx2.cell(row=1, column=5).border = lineborder
        wbx2.cell(row=1, column=6).border = lineborder
        wbx2.cell(row=1, column=8).border = lineborder
        wbx2.cell(row=2, column=8).border = lineborder
        wbx2.cell(row=1, column=9).border = lineborder
        wbx2.cell(row=2, column=9).border = lineborder
        wbx2.cell(row=3, column=9).border = lineborder
        wbx2.cell(row=4, column=9).border = lineborder
        wbx2.cell(row=5, column=9).border = lineborder
        wbx2.cell(row=6, column=9).border = lineborder
        wbx2.cell(row=7, column=9).border = lineborder
        wbx2.cell(row=1, column=10).border = lineborder
        wbx2.cell(row=2, column=10).border = lineborder
        wbx2.cell(row=3, column=10).border = lineborder
        wbx2.cell(row=4, column=10).border = lineborder
        wbx2.cell(row=5, column=10).border = lineborder
        wbx2.cell(row=6, column=10).border = lineborder
        wbx2.cell(row=7, column=10).border = lineborder
        
        wbx2.column_dimensions["A"].width = 18
        wbx2.column_dimensions["B"].width = 24
        wbx2.column_dimensions["C"].width = 12
        wbx2.column_dimensions["D"].width = 17
        wbx2.column_dimensions["E"].width = 12
        wbx2.column_dimensions["F"].width = 12
        wbx2.column_dimensions["H"].width = 10
        wbx2.column_dimensions["I"].width = 12
        wbx2.column_dimensions["J"].width = 18
        
        wbx2.row_dimensions[1].height = 36
        
        wbx3.cell(row=1, column=1).value = "実施時間"
        wbx3.cell(row=1, column=2).value = "名前"
        wbx3.cell(row=1, column=3).value = "会社名"
        wbx3.cell(row=1, column=4).value = "部署名"
        wbx3.cell(row=1, column=5).value = "Emailアドレス"        
        
        for x in range(1, 6):
            wbx3.cell(row=1, column=x).alignment = Alignment(horizontal = 'center', 
                                        vertical = 'center',
                                        wrap_text = False)
            wbx3.cell(row=1, column=x).fill = fill1
            wbx3.cell(row=1, column=x).font = boldfont
        
        for y in range(1, 6):
            wbx3.cell(row=1, column=y).border = lineborder
        
        wbx3.column_dimensions["A"].width = 20
        wbx3.column_dimensions["B"].width = 20
        wbx3.column_dimensions["C"].width = 24
        wbx3.column_dimensions["D"].width = 26
        wbx3.column_dimensions["E"].width = 30
            
        wbx.save(check_file_path)
        messagebox.showinfo("Message Box","[Email_Automation_template.xlsx]ファイルが作成されました。")
        
def button_function2():
    res = messagebox.askokcancel("Confirmation", "*実行する前に[Email_Automation_template.xlsx]ファイルが閉じていることを確認してください。" + "\n" + "*実行する前にOutlookを立ち上げてください。")
    check_dir_path = r"C:\email_automation"
    check_file_path = r"C:\email_automation\Email_Automation_template.xlsx"
    mailcount = 0
    
    if res == False:
        #sys.exit()
        pass    
    elif os.path.isdir(check_dir_path) and os.path.isfile(check_file_path):
        try:
            wb0 = px.load_workbook(check_file_path)
            ws0 = wb0.worksheets[0]
            wb0.save(check_file_path)
        except PermissionError :
            messagebox.showinfo("Message Box","*[Email_Automation_template.xlsx]ファイルを閉じてください。")
            pass
        else:    
            pypath = r"C:\email_automation"
            inputFile = pypath + "\\Email_Automation_template.xlsx"
            attachFile = pypath + "\\attachments\\"

            outlook = win32com.client.Dispatch("Outlook.Application")

            wb = px.load_workbook(inputFile)
            ws1 = wb.worksheets[0] #email contents
            ws2 = wb.worksheets[1] #list
            ws3 = wb.worksheets[2] #log

            max = ws2.max_row
            while True:
                if ws2.cell(row=max, column=1).value == None:
                    max -= 1
                else:
                    break

            fromaddress = ws1.cell(row=2, column=1).value
            fromaddress = fromaddress or ""
            cc = ws1.cell(row=2, column=3).value
            cc = cc or ""
            bcc = ws1.cell(row=5, column=3).value
            bcc = bcc or ""
            title = ws1.cell(row=5, column=1).value
            title = title or ""
            signature = ws1.cell(row=8, column=3).value
            signature = signature or ""


            A = ws2.cell(row=2, column=10).value
            A = A or ""
            B = ws2.cell(row=3, column=10).value
            B = B or ""
            C = ws2.cell(row=4, column=10).value
            C = C or ""
            D = ws2.cell(row=5, column=10).value
            D = D or ""
            E = ws2.cell(row=6, column=10).value
            E = E or ""
            F = ws2.cell(row=7, column=10).value
            F = F or ""

            attachFileA = attachFile + A
            attachFileB = attachFile + B
            attachFileC = attachFile + C
            attachFileD = attachFile + D
            attachFileE = attachFile + E
            attachFileF = attachFile + F

            for i in range(max-1):
                mail = outlook.CreateItem(0)
                if ws2.cell(row=i+2, column=6).value == "#":
                    
                    name = ws2.cell(row=i+2, column=1).value
                    name = name or ""
                    company = ws2.cell(row=i+2, column=2).value
                    company = company or ""
                    section = ws2.cell(row=i+2, column=3).value
                    section = section or ""
                    to = ws2.cell(row=i+2, column=4).value
                    to = to or ""
                    
                    mail_body = ws1.cell(row=8, column=1).value
                    mail_body = mail_body or ""
                    mail_body = company + "\n" + section + "\n" + name + "　様" + "\n" + "\n" + mail_body + "\n" + signature

                    mail.bodyFormat = 1 # 1:plain text, 2:HTML, 3:Rich Text 
                    mail.Sender = fromaddress
                    mail.To = to 
                    mail.CC = cc
                    mail.bcc = bcc
                    mail.Subject = title
                    mail.Body = mail_body 

                    if ws2.cell(row=i+2, column=5).value == "A" and A != "":
                        if not os.path.isfile(attachFileA):
                            messagebox.showinfo("Message Box", A + "という名前のファイルは存在しません。 [attachments]フォルダの中身を確認してください。メールは添付の無い状態で作成されます。")
                            pass
                        else:
                            mail.Attachments.Add(attachFileA)
                    elif ws2.cell(row=i+2, column=5).value == "B" and B != "":
                        if not os.path.isfile(attachFileB):
                            messagebox.showinfo("Message Box", B + "という名前のファイルは存在しません。 [attachments]フォルダの中身を確認してください。メールは添付の無い状態で作成されます。")
                            pass
                        else:
                            mail.Attachments.Add(attachFileB)
                    elif ws2.cell(row=i+2, column=5).value == "C" and C != "":
                        if not os.path.isfile(attachFileC):
                            messagebox.showinfo("Message Box", C + "という名前のファイルは存在しません。 [attachments]フォルダの中身を確認してください。メールは添付の無い状態で作成されます。")
                            pass
                        else:
                            mail.Attachments.Add(attachFileC)
                    elif ws2.cell(row=i+2, column=5).value == "D" and D != "":
                        if not os.path.isfile(attachFileD):
                            messagebox.showinfo("Message Box", D + "という名前のファイルは存在しません。 [attachments]フォルダの中身を確認してください。メールは添付の無い状態で作成されます。")
                            pass
                        else:
                            mail.Attachments.Add(attachFileD)
                    elif ws2.cell(row=i+2, column=5).value == "E" and E != "":
                        if not os.path.isfile(attachFileE):
                            messagebox.showinfo("Message Box", E + "という名前のファイルは存在しません。 [attachments]フォルダの中身を確認してください。メールは添付の無い状態で作成されます。")
                            pass
                        else:
                            mail.Attachments.Add(attachFileE)
                    elif ws2.cell(row=i+2, column=5).value == "F" and F != "":
                        if not os.path.isfile(attachFileF):
                            messagebox.showinfo("Message Box", F + "という名前のファイルは存在しません。 [attachments]フォルダの中身を確認してください。メールは添付の無い状態で作成されます。")
                            pass
                        else:
                            mail.Attachments.Add(attachFileF)
                    else:
                        pass
                    mail.display(True)
                    mailcount += 1
                    
                    max1 = ws3.max_row
                    while True:
                        if ws3.cell(row=max1, column=1).value == None:
                            max1 -= 1
                        else:
                            break
                    
                    ws2.cell(row=i+2, column=6).value = "Done"
                    ws3.cell(row=max1+1, column=1).value = datetime.datetime.now()
                    ws3.cell(row=max1+1, column=2).value = name
                    ws3.cell(row=max1+1, column=3).value = company
                    ws3.cell(row=max1+1, column=4).value = section
                    ws3.cell(row=max1+1, column=5).value = to

                    wb.save(inputFile)
                    
                    mail = None
                    name = None
                    company = None
                    section = None
                    to = None
                    mail_body = None
            if mailcount == 0:
                messagebox.showinfo("Message Box","メール送信対象のレコードがありません。")
            else:
                messagebox.showinfo("Message Box","完了!!")
                
    else:
        messagebox.showinfo("Message Box","ステップ[1] & [2]を実行したか確認してください。")
            
window=tkinter.Tk()
window.title("Email Automation")
window.geometry("500x200+350+150")#first 2digits => width of window, 3rd & 4th digits => place of window(x, y)
window.resizable(0,0)#unable to resize the window

label=tkinter.Label(window, text="[1] フォルダ作成 (事前環境準備)", font=("MSゴシック", "13"))
label.place(x=20, y=20)

button=tkinter.Button(window, text="フォルダ", font=("", 11), 
                    bg='#7b68ee', fg='linen', width=11, 
                    command=lambda:[button_function0()])
button.place(x=370, y=20)

label=tkinter.Label(window, text="[2] テンプレートファイル作成", font=("MSゴシック", "13"))
label.place(x=20, y=60)

button=tkinter.Button(window, text="テンプレート", font=("", 11), 
                    bg='#7b68ee', fg='linen', width=11, 
                    command=lambda:[button_function1()])
button.place(x=370, y=60)

label=tkinter.Label(window, text="[3] メール作成", font=("MSゴシック", "13"))
label.place(x=20, y=100)

button=tkinter.Button(window, text="メール", font=("", 11), 
                    bg='#7b68ee', fg='linen', width=11, 
                    command=lambda:[button_function2()])
button.place(x=370, y=100)

label=tkinter.Label(window, text="**事前に[メール自動化マニュアル.pdf]を確認してください。", font=("MSゴシック", "10"))
label.place(x=20, y=150)

#============================
window.mainloop()     
        




